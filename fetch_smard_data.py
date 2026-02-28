#!/usr/bin/env python3
"""
fetch_smard_data.py
═══════════════════════════════════════════════════════════════════════════════
Downloads real EPEX Spot day-ahead prices (15-min resolution) from SMARD.de.
Free, no login, no API key required.  CC BY 4.0 licence.

Usage:
    python fetch_smard_data.py                   # fetches 2022-2024 (default)
    python fetch_smard_data.py --years 2021 2024  # custom range
    python fetch_smard_data.py --test             # fetch just 1 month to test

Output:
    data/smard_prices_raw.csv        — raw 15-min EPEX spot (EUR/MWh)
    data/smard_prices_processed.csv  — all-in price (EUR/kWh) ready for model
    data/v2g_params.xlsx             — updated with real SeasonalPrices sheet

SMARD API (bundesAPI/smard-api, CC BY 4.0):
    Index:     GET /app/chart_data/{filter}/{region}/index_{resolution}.json
    Series:    GET /app/chart_data/{filter}/{region}/{filter}_{region}_{resolution}_{timestamp}.json

    filter 4169 = Day-ahead price (EUR/MWh), Germany/Luxembourg, 15-min
    filter 4170 = Day-ahead price (EUR/MWh), DE-AT-LU (pre-Oct 2018)

═══════════════════════════════════════════════════════════════════════════════
"""

from __future__ import annotations
import argparse
import json
import time
import warnings
from datetime import datetime, timezone, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    import urllib.request

# ── SMARD API constants ────────────────────────────────────────────────────────
BASE_URL   = "https://www.smard.de/app/chart_data"
FILTER_ID  = 4169          # Day-ahead price, DE-LU, quarterhour
REGION     = "DE-LU"
RESOLUTION = "quarterhour"

# BNetzA regulated fixed costs 2024 (EUR/kWh) — added to spot price
# Source: Bundesnetzagentur Monitoringbericht 2024
FIXED_COSTS = {
    "network_fee":    0.0663,   # Netzentgelt (avg. German commercial depot)
    "concession":     0.01992,  # Konzessionsabgabe
    "offshore_levy":  0.00816,  # Offshore-Netzumlage
    "chp_levy":       0.00277,  # KWKG-Umlage
    "electricity_tax":0.0205,   # Stromsteuer (§ 3 StromStG)
    "nev19":          0.01558,  # NEV-19-Umlage
}
FIXED_NET = sum(FIXED_COSTS.values())
VAT       = 0.19

# FCR/aFRR balancing premiums (EUR/kWh) — Agora Verkehrswende 2025, Fig. 4
FCR_PREMIUM_PEAK    = 0.132   # 16-20h weekdays
FCR_PREMIUM_MORNING = 0.040   # 07-09h

Path("data").mkdir(exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  HTTP helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _get(url: str, retries: int = 3, delay: float = 1.0) -> dict:
    """GET JSON with retry logic."""
    for attempt in range(retries):
        try:
            if HAS_REQUESTS:
                r = requests.get(url, timeout=20,
                                 headers={"Accept": "application/json"})
                r.raise_for_status()
                return r.json()
            else:
                req = urllib.request.Request(url,
                    headers={"Accept": "application/json",
                             "User-Agent": "V2G-thesis-fetcher/1.0"})
                with urllib.request.urlopen(req, timeout=20) as resp:
                    return json.loads(resp.read())
        except Exception as e:
            if attempt < retries - 1:
                print(f"    Retry {attempt+1}/{retries} after error: {e}")
                time.sleep(delay * (attempt + 1))
            else:
                raise
    return {}


# ═══════════════════════════════════════════════════════════════════════════════
#  SMARD API calls
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_available_timestamps() -> list[int]:
    """Fetch list of all available weekly timestamps from SMARD."""
    url = f"{BASE_URL}/{FILTER_ID}/{REGION}/index_{RESOLUTION}.json"
    print(f"  Fetching timestamp index from SMARD...")
    data = _get(url)
    timestamps = data.get("timestamps", [])
    print(f"  Found {len(timestamps)} weekly chunks available "
          f"(earliest: {_ts_to_dt(timestamps[0])}, "
          f"latest: {_ts_to_dt(timestamps[-1])})")
    return timestamps


def fetch_week_series(timestamp_ms: int) -> list[list]:
    """
    Fetch one week of 15-min price data for a given SMARD timestamp.
    Returns list of [timestamp_ms, price_EUR_per_MWh] pairs.
    Null values are returned as-is (handled later).
    """
    url = (f"{BASE_URL}/{FILTER_ID}/{REGION}/"
           f"{FILTER_ID}_{REGION}_{RESOLUTION}_{timestamp_ms}.json")
    data = _get(url)
    return data.get("series", [])


def _ts_to_dt(ts_ms: int) -> str:
    """Convert SMARD millisecond UTC timestamp to readable string."""
    return datetime.fromtimestamp(ts_ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d")


# ═══════════════════════════════════════════════════════════════════════════════
#  Download and assemble price series
# ═══════════════════════════════════════════════════════════════════════════════

def download_prices(year_start: int, year_end: int, test_mode: bool = False) -> pd.DataFrame:
    """
    Download all 15-min EPEX spot prices between year_start and year_end.

    Returns DataFrame with columns:
        timestamp_utc, datetime_local, price_EUR_MWh
    """
    all_timestamps = fetch_available_timestamps()

    # Filter to requested year range
    dt_start = datetime(year_start, 1, 1, tzinfo=timezone.utc)
    dt_end   = datetime(year_end, 12, 31, 23, 59, tzinfo=timezone.utc)

    filtered = [ts for ts in all_timestamps
                if dt_start.timestamp() * 1000 <= ts <= dt_end.timestamp() * 1000]

    if test_mode:
        filtered = filtered[:4]   # just 4 weeks ~1 month
        print(f"  TEST MODE: downloading only {len(filtered)} weeks")

    print(f"  Downloading {len(filtered)} weekly chunks "
          f"({year_start}–{year_end})...")

    rows = []
    for i, ts_ms in enumerate(filtered):
        week_dt = _ts_to_dt(ts_ms)
        print(f"    [{i+1:3d}/{len(filtered)}] week of {week_dt}", end="\r")
        try:
            series = fetch_week_series(ts_ms)
            for point in series:
                if len(point) == 2 and point[1] is not None:
                    rows.append({"timestamp_ms": point[0],
                                 "price_EUR_MWh": float(point[1])})
            time.sleep(0.15)   # polite rate limit ~7 req/s
        except Exception as e:
            print(f"\n    WARNING: could not fetch week {week_dt}: {e}")
            continue

    print(f"\n  Downloaded {len(rows):,} 15-min price points")

    df = pd.DataFrame(rows).drop_duplicates("timestamp_ms").sort_values("timestamp_ms")
    df["datetime_utc"] = pd.to_datetime(df["timestamp_ms"], unit="ms", utc=True)
    df["datetime_cet"] = df["datetime_utc"].dt.tz_convert("Europe/Berlin")
    df.set_index("datetime_cet", inplace=True)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
#  Feature engineering for the forecasting model
# ═══════════════════════════════════════════════════════════════════════════════

def engineer_features(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add all features needed for the price forecasting model.

    Lag features (Liu 2023; Lago et al. 2021):
        - Recent: t-1 .. t-4  (last hour in 15-min slots)
        - Same time yesterday: t-96
        - Same time last week: t-672
        - Rolling stats: 4h mean, 24h mean, 24h std

    Calendar features:
        - hour_sin/cos, dow_sin/cos, month_sin/cos (cyclical encoding)
        - is_weekend, is_holiday (DE public holidays)
    """
    df = df.copy()
    p = df["price_EUR_MWh"]

    # Lag features
    for lag in [1, 2, 3, 4, 8, 12, 24, 48, 96, 192, 672]:
        df[f"lag_{lag}"] = p.shift(lag)

    # Rolling statistics
    df["roll_4h_mean"]  = p.shift(1).rolling(16).mean()
    df["roll_24h_mean"] = p.shift(1).rolling(96).mean()
    df["roll_24h_std"]  = p.shift(1).rolling(96).std()
    df["roll_7d_mean"]  = p.shift(1).rolling(672).mean()

    # Calendar features (cyclical encoding avoids discontinuities)
    idx = df.index
    hour_frac = idx.hour + idx.minute / 60
    df["hour_sin"] = np.sin(2 * np.pi * hour_frac / 24)
    df["hour_cos"] = np.cos(2 * np.pi * hour_frac / 24)
    df["dow_sin"]  = np.sin(2 * np.pi * idx.dayofweek / 7)
    df["dow_cos"]  = np.cos(2 * np.pi * idx.dayofweek / 7)
    df["month_sin"]= np.sin(2 * np.pi * idx.month / 12)
    df["month_cos"]= np.cos(2 * np.pi * idx.month / 12)
    df["is_weekend"]= (idx.dayofweek >= 5).astype(int)

    # German public holidays (fixed + approximated movable)
    # Fixed national holidays
    fixed_holidays = {(1,1),(5,1),(10,3),(10,31),(11,1),(12,25),(12,26)}
    # Approximate movable holidays by month/day ranges
    def is_holiday(dt):
        return (dt.month, dt.day) in fixed_holidays
    df["is_holiday"] = [int(is_holiday(d)) for d in idx]

    # Season (quarter-based for simplicity)
    df["season"] = ((idx.month % 12) // 3).astype(int)   # 0=winter,1=spring,2=summer,3=autumn

    return df


def add_all_in_price(df: pd.DataFrame) -> pd.DataFrame:
    """Convert EPEX spot (EUR/MWh) to all-in depot price (EUR/kWh)."""
    df = df.copy()
    spot_kwh = df["price_EUR_MWh"] / 1000.0

    # All-in buy price (spot + regulated costs + VAT)
    df["buy_EUR_kWh"] = (spot_kwh + FIXED_NET) * (1 + VAT)

    # V2G feed-in price = buy + FCR/aFRR premium
    hour = df.index.hour + df.index.minute / 60
    fcr  = np.where((hour >= 16) & (hour < 20), FCR_PREMIUM_PEAK, 0.0)
    afrr = np.where((hour >= 7)  & (hour < 9),  FCR_PREMIUM_MORNING, 0.0)
    df["v2g_EUR_kWh"] = df["buy_EUR_kWh"] + fcr + afrr

    return df


# ═══════════════════════════════════════════════════════════════════════════════
#  Save to CSV + update Excel
# ═══════════════════════════════════════════════════════════════════════════════

def save_raw(df: pd.DataFrame, path: str = "data/smard_prices_raw.csv"):
    df[["timestamp_ms", "price_EUR_MWh"]].to_csv(path)
    print(f"  Raw data saved → {path}  ({len(df):,} rows)")


def save_processed(df: pd.DataFrame, path: str = "data/smard_prices_processed.csv"):
    cols = ["price_EUR_MWh", "buy_EUR_kWh", "v2g_EUR_kWh"] + \
           [c for c in df.columns if c.startswith("lag_") or
            c.startswith("roll_") or c.startswith("hour_") or
            c.startswith("dow_") or c.startswith("month_") or
            c in ("is_weekend","is_holiday","season")]
    out = df[cols].dropna()
    out.to_csv(path)
    print(f"  Processed data saved → {path}  ({len(out):,} rows, "
          f"{out.columns.tolist()[:5]}...)")
    return out


def update_excel(df_raw: pd.DataFrame,
                 excel_path: str = "data/v2g_params.xlsx"):
    """
    Update the SeasonalPrices sheet in v2g_params.xlsx with real median
    winter and summer daily profiles computed from real SMARD data.
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        print("  WARNING: openpyxl not available — skipping Excel update")
        return

    df = add_all_in_price(df_raw.copy())
    df["hour"] = df.index.hour + df.index.minute / 60
    df["month"] = df.index.month

    # Winter: Dec/Jan/Feb  |  Summer: Jun/Jul/Aug
    winter = df[df["month"].isin([12, 1, 2])]
    summer = df[df["month"].isin([6, 7, 8])]

    def median_day_profile(subset, col):
        return subset.groupby("hour")[col].median().reindex(
            np.arange(0, 24, 0.25), method="ffill").values[:96]

    slots = np.arange(96)
    hours = slots * 0.25
    time_labels = [f"{int(h):02d}:{int((h%1)*60):02d}" for h in hours]

    w_spot = median_day_profile(winter, "price_EUR_MWh") / 1000
    s_spot = median_day_profile(summer, "price_EUR_MWh") / 1000
    w_buy  = median_day_profile(winter, "buy_EUR_kWh")
    s_buy  = median_day_profile(summer, "buy_EUR_kWh")
    w_v2g  = median_day_profile(winter, "v2g_EUR_kWh")
    s_v2g  = median_day_profile(summer, "v2g_EUR_kWh")

    seasonal_df = pd.DataFrame({
        "Slot":                slots,
        "Time":                time_labels,
        "Hour":                hours,
        "Winter_EPEX_EUR_kWh": np.round(w_spot, 4),
        "Summer_EPEX_EUR_kWh": np.round(s_spot, 4),
        "Winter_Buy_EUR_kWh":  np.round(w_buy, 4),
        "Summer_Buy_EUR_kWh":  np.round(s_buy, 4),
        "Winter_V2G_EUR_kWh":  np.round(w_v2g, 4),
        "Summer_V2G_EUR_kWh":  np.round(s_v2g, 4),
        "Source":              [f"SMARD.de real data median profile"] * 96,
    })

    if Path(excel_path).exists():
        # Load and replace the SeasonalPrices sheet
        wb = load_workbook(excel_path)
        if "SeasonalPrices" in wb.sheetnames:
            del wb["SeasonalPrices"]
        ws = wb.create_sheet("SeasonalPrices")
        ws.append(list(seasonal_df.columns))
        for row in seasonal_df.itertuples(index=False):
            ws.append(list(row))
        wb.save(excel_path)
        print(f"  SeasonalPrices sheet updated in {excel_path}")
    else:
        print(f"  WARNING: {excel_path} not found — run make_data.py first, then re-run this script")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Fetch SMARD.de EPEX price data")
    parser.add_argument("--years", nargs=2, type=int, default=[2022, 2024],
                        metavar=("START", "END"),
                        help="Year range (default: 2022 2024)")
    parser.add_argument("--test", action="store_true",
                        help="Test mode: fetch only ~1 month of data")
    parser.add_argument("--skip-download", action="store_true",
                        help="Skip download, just reprocess existing raw CSV")
    args = parser.parse_args()

    print("\n" + "="*65)
    print("  SMARD.de EPEX Price Fetcher")
    print("  Data: EPEX Spot day-ahead, Germany, 15-min resolution")
    print("  Licence: CC BY 4.0 (Bundesnetzagentur)")
    print("="*65)

    raw_path = "data/smard_prices_raw.csv"

    if args.skip_download and Path(raw_path).exists():
        print(f"\n  Loading existing raw data from {raw_path}...")
        df_raw = pd.read_csv(raw_path, index_col=0, parse_dates=True)
    else:
        print(f"\n  Downloading {args.years[0]}–{args.years[1]} price data...")
        df_raw = download_prices(args.years[0], args.years[1],
                                 test_mode=args.test)
        save_raw(df_raw)

    print(f"\n  Date range: {df_raw.index.min()} → {df_raw.index.max()}")
    print(f"  Price range: {df_raw['price_EUR_MWh'].min():.1f} – "
          f"{df_raw['price_EUR_MWh'].max():.1f} EUR/MWh")
    print(f"  Negative price hours: "
          f"{(df_raw['price_EUR_MWh'] < 0).sum():,} / {len(df_raw):,} slots")

    print("\n  Engineering features...")
    df_feat = engineer_features(df_raw)
    df_feat = add_all_in_price(df_feat)
    save_processed(df_feat)

    print("\n  Updating v2g_params.xlsx with real seasonal profiles...")
    update_excel(df_raw)

    n_years = (df_raw.index.max() - df_raw.index.min()).days / 365
    print(f"\n  Summary: {len(df_raw):,} observations, {n_years:.1f} years")
    print(f"  Mean buy price: EUR {(df_raw['price_EUR_MWh']/1000 + FIXED_NET).mean() * (1+VAT):.3f}/kWh")
    print("\n  Done. Next step: run  python train_forecaster.py\n")


if __name__ == "__main__":
    main()
